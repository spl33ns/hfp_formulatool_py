from __future__ import annotations

import hashlib
import logging
from collections import defaultdict
from pathlib import Path

from core.dnf import normalize_dnf, to_dnf
from core.formula_parser import FormulaParseError, parse_formula_with_literal_parser, parse_literal
from core.models import LiteralModel, SectionKey, SectionResult
from core.operator_config import OperatorConfig, OperatorConfigError, load_operator_config
from core.stages import Stage
from core.utils import ABBREVIATIONS, ensure_unique_sheet_name, sanitize_filename
from core.variable_mapping import (
    VariableMappingError,
    VariableMeta,
    load_variable_mapping_tsv,
    lookup_variable_meta,
)
from exporters.confluence import write_confluence_markdown
from exporters.docs_exporter import write_docs_files
from exporters.xlsx_exporter import write_section_sheet

logger = logging.getLogger(__name__)


class SectionFailure(Exception):
    pass


def _extra(stage: Stage, section: str | None) -> dict[str, str]:
    return {"stage": stage.value, "section": section or "-"}


def _mark_logged(exc: BaseException) -> None:
    try:
        setattr(exc, "_hfp_logged", True)
    except Exception:
        pass


def _is_logged(exc: BaseException) -> bool:
    return bool(getattr(exc, "_hfp_logged", False))


def _format_section_ref(
    env_objective: str | None,
    activity: str | None,
    dnsh_goal: str | None,
    formula_ids: str | None,
) -> str:
    def _clean(value: str | None) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("\r", " ").replace("\t", " ").strip()

    a = _clean(env_objective)
    c = _clean(activity)
    f = _clean(dnsh_goal)
    h = _clean(formula_ids)
    h_short = h if len(h) <= 80 else (h[:77] + "...")

    # The section key is (A,C,F,H). The fingerprint only depends on these fields.
    fingerprint_src = "|".join([a, c, f, h])
    fingerprint = hashlib.md5(fingerprint_src.encode("utf-8")).hexdigest()[:8]

    return f"A={a} | C={c} | F={f} | H={h_short} | id={fingerprint}"


def _run_stage(stage: Stage, section: str | None, func, expected_exceptions: tuple[type[BaseException], ...] = ()):
    logger.info("START", extra=_extra(stage, section))
    try:
        value = func()
    except expected_exceptions as exc:
        logger.error("FAILED: %s", exc, exc_info=True, extra=_extra(stage, section))
        _mark_logged(exc)
        raise
    except Exception as exc:
        logger.exception("FAILED: unexpected error", extra=_extra(stage, section))
        _mark_logged(exc)
        raise
    else:
        logger.info("OK", extra=_extra(stage, section))
        return value


def build_sheet_name(env_objective: str, section_number: str) -> str:
    if env_objective not in ABBREVIATIONS:
        raise SectionFailure(f"Unknown environmental objective: {env_objective}")
    abbrev = ABBREVIATIONS[env_objective]
    raw_name = f"{abbrev}_{section_number}"
    return raw_name


def parse_formula_ids(
    formula_ids: str,
    op_config: OperatorConfig,
    *,
    log_extra: dict[str, str] | None = None,
) -> tuple[list[list[LiteralModel]], list[LiteralModel]]:
    id_literals: list[LiteralModel] = []

    def id_literal_parser(raw_literal: str) -> LiteralModel:
        parsed = parse_literal(raw_literal, raw_literal, op_config=op_config)
        # Column H is the source of truth (IDs). Use IDs for display_name by default.
        literal = LiteralModel(id=parsed.id, display_name=parsed.id, op=parsed.op)
        id_literals.append(literal)
        return literal

    id_ast = parse_formula_with_literal_parser(formula_ids, id_literal_parser, op_config=op_config, log_extra=log_extra)

    dnf_clauses = to_dnf(id_ast)
    normalized = normalize_dnf(dnf_clauses)

    deduped_vars: dict[str, LiteralModel] = {}
    for lit in id_literals:
        if lit.id not in deduped_vars:
            deduped_vars[lit.id] = lit
    return normalized, list(deduped_vars.values())


def _output_variable_sort_key(literal: LiteralModel) -> tuple[int, str, str]:
    question_text = literal.question_text.strip()
    return (
        1 if not question_text else 0,
        question_text.casefold(),
        literal.id.casefold(),
    )


def _enrich_variable_with_mapping(literal: LiteralModel, mapping: dict[str, VariableMeta]) -> tuple[LiteralModel, bool]:
    meta = lookup_variable_meta(mapping, literal.id)
    if meta is None:
        return (
            LiteralModel(
                id=literal.id,
                display_name=literal.display_name,
                op=literal.op,
                technical_name="",
                question_text="",
            ),
            False,
        )
    return (
        LiteralModel(
            id=literal.id,
            display_name=literal.display_name,
            op=literal.op,
            technical_name=meta.technical_name,
            question_text=meta.question_text,
        ),
        True,
    )


def process_excel(
    input_path: Path,
    output_root: Path,
    max_rules: int,
    mapping_path: Path | None = None,
) -> dict[str, list[SectionResult]]:
    from openpyxl import Workbook, load_workbook

    logger.info("START processing", extra=_extra(Stage.RUN, None))
    logger.info("Input=%s Output=%s max_rules=%s", input_path, output_root, max_rules, extra=_extra(Stage.RUN, None))

    op_config = _run_stage(
        Stage.LOAD_OPERATOR_CONFIG,
        None,
        lambda: load_operator_config(log_extra=_extra(Stage.LOAD_OPERATOR_CONFIG, None)),
        expected_exceptions=(OperatorConfigError,),
    )

    variable_mapping: dict[str, VariableMeta] = {}
    mapping_loaded = False
    mapping_stage_extra = _extra(Stage.LOAD_VARIABLE_MAPPING, None)
    if mapping_path is None:
        logger.info("NO_MAPPING_CONFIGURED", extra=mapping_stage_extra)
    else:
        try:
            variable_mapping = _run_stage(
                Stage.LOAD_VARIABLE_MAPPING,
                None,
                lambda: load_variable_mapping_tsv(mapping_path, log_extra=mapping_stage_extra),
                expected_exceptions=(VariableMappingError,),
            )
            mapping_loaded = True
        except VariableMappingError:
            # Config-load failure is already logged with stacktrace; continue with ID-only fallback.
            pass

    workbook = _run_stage(
        Stage.LOAD_WORKBOOK,
        None,
        lambda: load_workbook(input_path, data_only=True),
    )
    sheet = workbook.active

    grouped: dict[tuple[str | None, str | None, str | None, str | None], list[dict[str, str | None]]] = defaultdict(list)
    def _group_rows() -> None:
        for row in sheet.iter_rows(min_row=2):
            values = {
                "A": row[0].value,
                "B": row[1].value,
                "C": row[2].value,
                "F": row[5].value,
                "G": row[6].value,
                "H": row[7].value,
                "I": row[8].value,
                "J": row[9].value if len(row) > 9 else None,
            }
            key = (values["A"], values["C"], values["F"], values["H"])
            grouped[key].append(values)

    _run_stage(Stage.GROUP_ROWS, None, _group_rows)
    logger.info("Grouped sections=%d", len(grouped), extra=_extra(Stage.GROUP_ROWS, None))

    activity_results: dict[str, list[SectionResult]] = defaultdict(list)
    missing_mapping_ids: list[str] = []
    missing_mapping_seen: set[str] = set()

    for key, rows in grouped.items():
        env_objective, activity, dnsh_goal, formula_ids = key
        first = rows[0]
        section_number = first["B"]
        type_label = first["G"]
        formula_display = first["I"]
        section_ref = _format_section_ref(
            str(env_objective) if env_objective is not None else None,
            str(activity) if activity is not None else None,
            str(dnsh_goal) if dnsh_goal is not None else None,
            str(formula_ids) if formula_ids is not None else None,
        )
        section_key = SectionKey(
            env_objective=str(env_objective) if env_objective is not None else None,
            section_number=str(section_number) if section_number is not None else None,
            activity=str(activity) if activity is not None else None,
            dnsh_goal=str(dnsh_goal) if dnsh_goal is not None else None,
            formula_ids=str(formula_ids) if formula_ids is not None else None,
            type_label=str(type_label) if type_label is not None else None,
        )
        activity_name = str(activity) if activity else "Unknown"
        try:
            def _validate_required_fields() -> None:
                if not env_objective or not activity or not dnsh_goal or not formula_ids:
                    raise SectionFailure("Missing required section identification fields")
                if not section_number:
                    raise SectionFailure("Missing section number")
                if not formula_display:
                    raise SectionFailure("Missing display formula")

            _run_stage(Stage.SECTION_VALIDATE, section_ref, _validate_required_fields, expected_exceptions=(SectionFailure,))

            sheet_name = _run_stage(
                Stage.SECTION_SHEET_NAME,
                section_ref,
                lambda: build_sheet_name(str(env_objective), str(section_number)),
                expected_exceptions=(SectionFailure,),
            )
            dnf_clauses, variables = _run_stage(
                Stage.SECTION_PARSE,
                section_ref,
                lambda: parse_formula_ids(
                    str(formula_ids),
                    op_config,
                    log_extra=_extra(Stage.SECTION_PARSE, section_ref),
                ),
                expected_exceptions=(SectionFailure, FormulaParseError, OperatorConfigError),
            )

            def _check_rule_limit() -> None:
                if len(dnf_clauses) > max_rules:
                    raise SectionFailure("DNF rule limit exceeded")

            _run_stage(Stage.SECTION_MAX_RULES, section_ref, _check_rule_limit, expected_exceptions=(SectionFailure,))

            enriched_variables: list[LiteralModel] = []
            for variable in variables:
                enriched, found = _enrich_variable_with_mapping(variable, variable_mapping if mapping_loaded else {})
                enriched_variables.append(enriched)
                if mapping_loaded and not found and variable.id not in missing_mapping_seen:
                    missing_mapping_seen.add(variable.id)
                    missing_mapping_ids.append(variable.id)

            result = SectionResult(
                key=section_key,
                sheet_name=sheet_name,
                formula_ids=str(formula_ids),
                formula_display=str(formula_display),
                variables=sorted(enriched_variables, key=_output_variable_sort_key),
                dnf=dnf_clauses,
                status="OK",
                error=None,
            )
        except (SectionFailure, FormulaParseError) as exc:
            result = SectionResult(
                key=section_key,
                sheet_name=None,
                formula_ids=str(formula_ids) if formula_ids else None,
                formula_display=str(formula_display) if formula_display else None,
                variables=[],
                dnf=[],
                status="FAILED",
                error=str(exc),
            )
        except Exception as exc:
            if not _is_logged(exc):
                logger.exception("FAILED: hard fail while processing section", extra=_extra(Stage.RUN, section_ref))
            result = SectionResult(
                key=section_key,
                sheet_name=None,
                formula_ids=str(formula_ids) if formula_ids else None,
                formula_display=str(formula_display) if formula_display else None,
                variables=[],
                dnf=[],
                status="FAILED",
                error=str(exc),
            )
        activity_results[activity_name].append(result)

    for activity, sections in activity_results.items():
        activity_ref = f"C={activity}"

        def _init_activity_folders() -> tuple[Path, Path]:
            activity_folder = output_root / sanitize_filename(activity)
            activity_folder.mkdir(parents=True, exist_ok=True)
            confluence_folder = activity_folder / "confluence"
            confluence_folder.mkdir(exist_ok=True)
            return activity_folder, confluence_folder

        activity_folder, confluence_folder = _run_stage(Stage.EXPORT_ACTIVITY_INIT, activity_ref, _init_activity_folders)

        workbook_out = Workbook()
        workbook_out.remove(workbook_out.active)
        used_sheet_names: set[str] = set()
        created_sheet = False
        for section in sections:
            if section.status != "OK":
                continue
            section_ref = _format_section_ref(
                section.key.env_objective,
                section.key.activity,
                section.key.dnsh_goal,
                section.key.formula_ids,
            )

            try:
                def _write_sheet() -> None:
                    assert section.sheet_name
                    unique_name = ensure_unique_sheet_name(section.sheet_name, used_sheet_names)
                    section.sheet_name = unique_name
                    write_section_sheet(workbook_out, section, activity)

                _run_stage(Stage.EXPORT_SECTION_SHEET, section_ref, _write_sheet)
                created_sheet = True

                _run_stage(
                    Stage.EXPORT_SECTION_CONFLUENCE,
                    section_ref,
                    lambda: write_confluence_markdown(confluence_folder, section, activity),
                )
            except Exception as exc:
                # Stage logger already captured stacktrace; reflect failure in result for docs/summary.
                section.status = "FAILED"
                section.error = f"Export failed: {exc}"
                continue

        if not created_sheet:
            def _create_placeholder_sheet() -> None:
                placeholder = workbook_out.create_sheet("No_Valid_Sections")
                placeholder["A1"] = "No valid sections were generated for this activity."

            _run_stage(Stage.EXPORT_ACTIVITY_PLACEHOLDER, activity_ref, _create_placeholder_sheet)

        workbook_path = activity_folder / f"{sanitize_filename(activity)}.xlsx"
        try:
            _run_stage(Stage.EXPORT_ACTIVITY_SAVE_WORKBOOK, activity_ref, lambda: workbook_out.save(workbook_path))
        except Exception:
            # Keep going: docs export may still be helpful.
            pass

        try:
            _run_stage(Stage.EXPORT_ACTIVITY_DOCS, activity_ref, lambda: write_docs_files(activity_folder, activity, sections))
        except Exception:
            pass

    if mapping_loaded and missing_mapping_ids:
        logger.warning(
            "MISSING_MAPPING: missing=%d sample=%s",
            len(missing_mapping_ids),
            missing_mapping_ids[:20],
            extra=_extra(Stage.RUN, None),
        )

    logger.info("DONE processing", extra=_extra(Stage.RUN, None))
    return activity_results
