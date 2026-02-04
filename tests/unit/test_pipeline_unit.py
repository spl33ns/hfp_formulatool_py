from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from core.pipeline import process_excel


def _write_workbook(path: Path, rows: list[dict[str, str | None]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Env Objective",
            "Section",
            "Activity",
            "D",
            "E",
            "DNSH Goal",
            "Type",
            "Formula IDs",
            "Formula Display",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.get("A"),
                row.get("B"),
                row.get("C"),
                None,
                None,
                row.get("F"),
                row.get("G"),
                row.get("H"),
                row.get("I"),
            ]
        )
    wb.save(path)


def test_process_excel_happy_path_creates_outputs(tmp_path):
    input_path = tmp_path / "input.xlsx"
    _write_workbook(
        input_path,
        [
            {
                "A": "Climate Change Mitigation",
                "B": "3.5",
                "C": "Activity 1",
                "F": "Goal 1",
                "G": "Type A",
                "H": "A=1 AND B=1",
                "I": "A=1 AND B=1",
            }
        ],
    )

    output_root = tmp_path / "out"
    results = process_excel(input_path, output_root, 2000)

    assert "Activity 1" in results
    section = results["Activity 1"][0]
    assert section.status == "OK"
    assert section.sheet_name == "CCM_3.5"

    workbook_path = output_root / "Activity_1" / "Activity_1.xlsx"
    assert workbook_path.exists()

    wb = load_workbook(workbook_path)
    assert "CCM_3.5" in wb.sheetnames
    assert "No_Valid_Sections" not in wb.sheetnames


def test_process_excel_all_failed_creates_placeholder_sheet(tmp_path):
    input_path = tmp_path / "input.xlsx"
    _write_workbook(
        input_path,
        [
            {
                "A": "Climate Change Mitigation",
                "B": "3.5",
                "C": "Activity 1",
                "F": "Goal 1",
                "G": "Type A",
                "H": "A=1",
                "I": None,
            }
        ],
    )

    output_root = tmp_path / "out"
    results = process_excel(input_path, output_root, 2000)

    section = results["Activity 1"][0]
    assert section.status == "FAILED"
    assert section.sheet_name is None
    assert section.error and "Missing display formula" in section.error

    workbook_path = output_root / "Activity_1" / "Activity_1.xlsx"
    wb = load_workbook(workbook_path)
    assert wb.sheetnames == ["No_Valid_Sections"]


def test_process_excel_enforces_max_rules(tmp_path):
    input_path = tmp_path / "input.xlsx"
    _write_workbook(
        input_path,
        [
            {
                "A": "Climate Change Mitigation",
                "B": "3.5",
                "C": "Activity 1",
                "F": "Goal 1",
                "G": "Type A",
                "H": "A OR B OR C",
                "I": "A OR B OR C",
            }
        ],
    )

    results = process_excel(input_path, tmp_path / "out", 2)
    section = results["Activity 1"][0]

    assert section.status == "FAILED"
    assert section.error and "DNF rule limit exceeded" in section.error


def test_process_excel_structure_mismatch_fails(tmp_path):
    input_path = tmp_path / "input.xlsx"
    _write_workbook(
        input_path,
        [
            {
                "A": "Climate Change Mitigation",
                "B": "3.5",
                "C": "Activity 1",
                "F": "Goal 1",
                "G": "Type A",
                "H": "A AND B",
                "I": "A OR B",
            }
        ],
    )

    results = process_excel(input_path, tmp_path / "out", 2000)
    section = results["Activity 1"][0]

    assert section.status == "FAILED"
    assert section.error and "Formula structure mismatch between H and I" in section.error


def test_process_excel_duplicate_display_names_fails(tmp_path):
    input_path = tmp_path / "input.xlsx"
    _write_workbook(
        input_path,
        [
            {
                "A": "Climate Change Mitigation",
                "B": "3.5",
                "C": "Activity 1",
                "F": "Goal 1",
                "G": "Type A",
                "H": "A AND B",
                "I": "X AND X",
            }
        ],
    )

    results = process_excel(input_path, tmp_path / "out", 2000)
    section = results["Activity 1"][0]

    assert section.status == "FAILED"
    assert section.error and "Duplicate display name for different IDs" in section.error
