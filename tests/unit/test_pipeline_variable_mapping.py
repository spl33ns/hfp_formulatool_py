from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.pipeline import process_excel


def _write_input_workbook(path: Path, formula_ids: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Env Objective",
            "Section",
            "Activity",
            "D",
            "E",
            "DNSH Goal",
            "Type",
            "Formula IDs",
            "Formula Display",
        ]
    )
    ws.append(
        [
            "Climate Change Mitigation",
            "3.5",
            "Activity 1",
            None,
            None,
            "Goal 1",
            "Type A",
            formula_ids,
            "Display Formula",
        ]
    )
    wb.save(path)


def _write_mapping_file(path: Path, rows: list[tuple[str, str, str]]) -> None:
    lines = ["\t".join(["ID", "x", "Technical name", "x", "x", "x", "x", "Question text"])]
    for var_id, technical_name, question_text in rows:
        lines.append("\t".join([var_id, "x", technical_name, "x", "x", "x", "x", question_text]))
    path.write_text("\n".join(lines), encoding="utf-8")


def test_sorting_by_question_text_then_id(tmp_path):
    input_path = tmp_path / "input.xlsx"
    mapping_path = tmp_path / "mapping.csv"
    _write_input_workbook(input_path, "A2|A1|B1|C1")
    _write_mapping_file(
        mapping_path,
        [
            ("A2", "Tech A2", "Alpha"),
            ("A1", "Tech A1", "alpha"),
            ("B1", "Tech B1", "beta"),
            ("C1", "Tech C1", ""),
        ],
    )

    results = process_excel(input_path, tmp_path / "out", 2000, mapping_path=mapping_path)
    section = results["Activity 1"][0]

    assert section.status == "OK"
    assert [var.id for var in section.variables] == ["A1", "A2", "B1", "C1"]


def test_missing_mapping_emits_warning_and_blank_fields(tmp_path, caplog):
    input_path = tmp_path / "input.xlsx"
    mapping_path = tmp_path / "mapping.csv"
    _write_input_workbook(input_path, "A1|B1")
    _write_mapping_file(mapping_path, [("A1", "Tech A1", "Question A1")])

    caplog.set_level(logging.WARNING)
    results = process_excel(input_path, tmp_path / "out", 2000, mapping_path=mapping_path)
    section = results["Activity 1"][0]

    var_by_id = {var.id: var for var in section.variables}
    assert var_by_id["A1"].technical_name == "Tech A1"
    assert var_by_id["A1"].question_text == "Question A1"
    assert var_by_id["B1"].technical_name == ""
    assert var_by_id["B1"].question_text == ""
    assert any("MISSING_MAPPING" in record.message for record in caplog.records)
    assert any("B1" in record.message for record in caplog.records if "MISSING_MAPPING" in record.message)


def test_no_mapping_file_falls_back_to_id_only_for_excel_and_markdown(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_root = tmp_path / "out"
    _write_input_workbook(input_path, "A1|B1")

    results = process_excel(input_path, output_root, 2000)
    section = results["Activity 1"][0]
    assert section.status == "OK"
    assert all(var.technical_name == "" and var.question_text == "" for var in section.variables)

    workbook_path = output_root / "Activity_1" / "Activity_1.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook[section.sheet_name]
    assert sheet.cell(row=7, column=1).value == "ID"
    assert sheet.cell(row=7, column=2).value == "Technical name"
    assert sheet.cell(row=7, column=3).value == "Question text"

    markdown_path = output_root / "Activity_1" / "confluence" / f"{section.sheet_name}.md"
    markdown = markdown_path.read_text(encoding="utf-8")
    assert "| ID | Technical name | Question text |" in markdown
    assert "| A1 |  |  |" in markdown


def test_markdown_includes_technical_name_and_question_text_when_mapping_present(tmp_path):
    input_path = tmp_path / "input.xlsx"
    mapping_path = tmp_path / "mapping.csv"
    output_root = tmp_path / "out"
    _write_input_workbook(input_path, "A1|B1")
    _write_mapping_file(
        mapping_path,
        [
            ("A1", "Tech A1", "Question A1"),
            ("B1", "Tech B1", "Question B1"),
        ],
    )

    results = process_excel(input_path, output_root, 2000, mapping_path=mapping_path)
    section = results["Activity 1"][0]
    markdown_path = output_root / "Activity_1" / "confluence" / f"{section.sheet_name}.md"
    markdown = markdown_path.read_text(encoding="utf-8")

    assert "| ID | Technical name | Question text |" in markdown
    assert "Tech A1" in markdown
    assert "Question A1" in markdown


def test_lookup_with_suffix_id_uses_normalized_mapping_without_missing_warning(tmp_path, caplog):
    input_path = tmp_path / "input.xlsx"
    mapping_path = tmp_path / "mapping.csv"
    _write_input_workbook(input_path, "ePBN113509_1")
    _write_mapping_file(mapping_path, [("ePBN113509", "Tech Base", "Question Base")])

    caplog.set_level(logging.WARNING)
    results = process_excel(input_path, tmp_path / "out", 2000, mapping_path=mapping_path)
    section = results["Activity 1"][0]
    variable = section.variables[0]

    assert section.status == "OK"
    assert variable.id == "ePBN113509_1"
    assert variable.technical_name == "Tech Base"
    assert variable.question_text == "Question Base"
    assert not any("MISSING_MAPPING" in record.message for record in caplog.records)
